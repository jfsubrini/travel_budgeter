# -*- coding: utf-8 -*-
"""
Collecting the data from the Currencies array stored in a .xlsx file.
Creating the currencies dictionary, exported into the currencies_data.json file.
"""

import json
import openpyxl


WB = openpyxl.load_workbook("currencies.xlsx")
SHEET = WB["Currencies"]


def create_currencies_data():
    """Create the currencies_data with all the information, by row (countries),
    stoked into the cells from each column."""
    if SHEET["A1"].value != "country":
        raise ValueError(
            "Le tableau a subi une translation ! Elle doit commencer à la cellule A1."
        )
    matrix_data = {}
    dict_data = {}
    for row in range(2, SHEET.max_row + 1):
        for column in range(1, SHEET.max_column + 1):
            cell = SHEET.cell(row=row, column=column)
            column_name = SHEET.cell(row=1, column=column).value
            cell_value = cell.value
            dict_data[column_name] = cell_value
        row_name = SHEET.cell(row=row, column=1).value
        matrix_data[row_name] = dict_data
        dict_data = {}
    return matrix_data


if __name__ == "__main__":
    MATRIX_DATA = create_currencies_data()
    with open("currencies.json", "w") as file:
        json.dump(MATRIX_DATA, file, indent=4)
